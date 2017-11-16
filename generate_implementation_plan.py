# -*- coding: utf-8 -*-
from docxtpl import DocxTemplate, InlineImage
# for height and width you have to use millimeters (Mm), inches or points(Pt) class :
from docx.shared import Mm, Inches, Pt
from datetime import date
from openpyxl import load_workbook
import re
import textfsm
import os
import glob
import netmiko
import argparse
import pprint


TEMPLATE = 'Implemetation Plan Temlate.docx'
TEMPLATE_DD = './DD Template.docx'
SITES_LIST_FILE = './input_sites.txt'
SITES_ADDRESS = 'C:\\Users\\g84086619\\Desktop\\Armen\Sites\\SiteSurvey - Copy\\SiteSurvey - Copy\\ISDP (Total list of equipment new 2 0).xlsx'
IP_TABLE_FILE = 'C:\\Users\\g84086619\\Desktop\\Armen\\IP\\new_MGMT(15-11-2017).xlsx'
TEMPLATE_CISCO = 'C:\\Users\\g84086619\\Documents\\Devman\\Armentel\\Temlates\\'
SITES_SOFT_CUR = 'C:\\Users\\g84086619\\Desktop\\Armen\\config\\Current\\splitted-output\\splitted-output\\'
SITES_SOFT_TAR = 'C:\\Users\\g84086619\\Desktop\\Armen\\config\\Target\\15092017\\15092017\\'
SITES_FOLDER = 'C:\\Users\\g84086619\\Desktop\\Armen\\Sites\\SiteSurvey - Copy\\SiteSurvey - Copy\\'
LAYOUT_FOLDER = 'C:\\Users\\g84086619\\Documents\\Devman\\Armentel\\Layout\\'
TARGET_NETWORK_IMG = 'C:\\Users\\g84086619\\Desktop\\Armen\\topology\\fullnet.jpg'
LAYOUT_DESIGN = 'C:\\Users\\g84086619\\Desktop\\Armen\\LLD\\Armenia_Veon_Solution_LLD_v0.9_07102017.xlsx'



def generate_template(tpl, site, site_information):
    print('generate_template',site)
    site_address = get_site_address(site)
    site_device = get_site_device(site)
    search_city_name = re.search('[-,_](?P<city>[A-Z][a-z]+)', site)
    if search_city_name:
        city = search_city_name.group('city')
    else:
        city = 'No Information'
    print('generate_template city', city)
    software = get_software(site, site_information)
    context = {
        'layout': InlineImage(tpl, LAYOUT_FOLDER+site_device['target_d']+'.jpg', width=Mm(100)),
        'engineer': ' Anton Gorynya',
        'engineer_id': 'g84086619',
        'engineer_tel': '+7 931 206-67-90',
        'site_location': site_address['adress'],
        'gps': site_address['coordinates'],
        'target_network_img': InlineImage(tpl, TARGET_NETWORK_IMG, width=Mm(80)),
        'cisco_model': site_device['current_d'],
        'huawei_model': site_device['target_d'],
        'site': site,
        'city': city,
        'date_of_swap': site_information['Installation date'].strftime("%m/%d/%Y"),
        'cisco_soft': software['current_d'],
        'huawei_version': software['target_d'],
        'date': '04/11/2017',

        'hostname': site_information['old_hostname'],
        'site_plan': 'No Information'
    }
    return context


def update_context_from_ssh_data(site_information, context_dict):
    loopback = site_information['loopback'].split()[0]
    ssh = connect_to_device(loopback, 'huawei')
    print('extract_huawei_description')
    huawei_int_description = send_show_command(ssh, 'display  interface description ').split('\n')
    huawei_int_description = [
        [interface_des[:interface_des.find('  ')], interface_des[interface_des.find('  ') + 1:].strip()] for
        interface_des in huawei_int_description]
    print('send display  interface')
    dis_int = send_show_command(ssh, 'display  interface')
    print('extract media type')
    media_type = get_media_type(dis_int)
    print('create interface table')
    interface_table = create_interface_table(cisco_int_description, huawei_int_description, media_type)
    print('get vlan')
    vlan_content = generate_vlan_table_content(ssh)
    print('get configuration {}'.format(site_information['old_hostname']))
    configration = get_configuration(site_information['old_hostname'], ssh)
    if configration:
        new_config = configration['new_config']
        new_config = new_config.replace('&', 'and')
        new_config = new_config.replace('<', '')
        new_config = new_config.replace('>', '')
        new_config = new_config.replace('\n', '<w:br/>')
        old_config = configration['old_config'].replace('\n', '<w:br/>')
    else:
        new_config = 'No Information'
        old_config = 'No Information'
    context_dict.update({'cisco_configuration': old_config})
    context_dict.update({'huawei_configuration': new_config})
    print('generate_context')
    context_dict.update({'vlan_table': vlan_content})
    context_dict.update({'interface_table': interface_table})
    print('send display  license')
    h_license = send_show_command(ssh, 'display  license').replace('\n', '<w:br/>')
    context_dict.update({'license': h_license})
    print('send display  device')
    device_status = send_show_command(ssh, 'display  device ').replace('\n', '<w:br/>')
    context_dict.update({'device_status': device_status})
    print('display  interface brief')
    dis_interfaces_b = send_show_command(ssh, 'display  interface brief').replace('\n', '<w:br/>')
    context_dict.update({'dis_interfaces_brief': dis_interfaces_b})
    disconnect(ssh)
    return context_dict


def get_sites_list():
    with open(SITES_LIST_FILE) as sites_list:
        sites_list = sites_list.read().split()
    return sites_list


def genetate_implementation_design(context, tpl, doc_type):
    with open("{}.txt".format(context['site']), 'w') as out_f:
        print('context saved')
        out_f.write(str(context))
    tpl.render(context)
    tpl.save("Out/{} {}.docx".format(context['site'], doc_type))
    print('temlate saved')


def get_site_address(site):
    host = re.search('[-,_][a-z]*\.?(?P<host>[A-Z]*[a-z]*\d*)', site)
    host = host.group('host')
    wb = load_workbook(SITES_ADDRESS)
    second_sheet = wb.get_sheet_names()[1]
    worksheet = wb.get_sheet_by_name(second_sheet)
    for row in worksheet.iter_rows():
        if host == ''.join(row[2].value.split()):
            sites_address = {'adress': 'Province '
                                     + row[7].value
                                     + ' ' + row[4].value,
                           'coordinates': 'latitude:{} longitude:{}'.format(row[5].value, row[6].value)}
            return sites_address


def get_site_device(site):
    host = re.search('[-,_][a-z]*\.?(?P<host>[A-Z]*[a-z]*\d*)', site)
    host = host.group('host')
    print('host: {}'.format(host))
    wb = load_workbook(SITES_ADDRESS)
    first_sheet = wb.get_sheet_names()[0]
    worksheet = wb.get_sheet_by_name(first_sheet)
    for row in worksheet.iter_rows():
        if row[1].value != None and host == ''.join(row[1].value.split()):
            site_device = {'current_d': row[3].value, 'target_d': row[4].value}
            print(' current d: {}\n target d: {})'.format(site_device['current_d'], site_device['target_d']))
            return site_device


def get_sites_information():
    sites_information = {}
    wb = load_workbook(IP_TABLE_FILE)
    first_sheet = wb.get_sheet_names()[0]
    worksheet = wb.get_sheet_by_name(first_sheet)

    for row in worksheet.iter_rows():
        sites_information.update({row[2].value:
                                      {'old_hostname': row[3].value,
                                       'Installation date': row[0].value,
                                       'loopback': row[4].value,
                                       'MGMT_IP': row[5].value,
                                       'Gateway': row[6].value,
                                       'Cisco_first_int': row[7].value,
                                       'Cisco_second_int': row[8].value}})
    return sites_information


def get_software(site, site_information):
    if site in sites_information:
        print('hostname', site_information['old_hostname'])
        print ('Loopback:', site_information['loopback'])
        with open('{}{}.version'.format(SITES_SOFT_CUR, site_information['old_hostname'])) as version_file:
            version_file = version_file.read()
        with open(TEMPLATE_CISCO+'cisco_ios_show_version.template') as template:
            re_table = textfsm.TextFSM(template)
            result = re_table.ParseText(version_file)
            current_d = {'cisco_model': result[0][5],'cisco_soft': result[0][0]}
            return {'current_d':result[0][0], 'target_d': 'V800R008C10SPC500'}


def get_topology(ssh):
    print ('gather  interfaces  status  ')
    int_status = send_show_command(ssh,
                                   'show interfaces  status ')
    template_path = TEMPLATE_CISCO+'cisco_ios_show_interfaces_status.template'
    with open(template_path) as template:
        re_table = textfsm.TextFSM(template)
        result = re_table.ParseText(int_status)
    header = ['Port', 'Name', 'Status', 'Vlan or Routed', 'Duplex', 'Speed', 'Type']
    active_ports = {'header': header, 'content': result}
    return active_ports

def get_topology_from_file(site):
    host = re.search('[-,_](?P<host>[a-z]*\.?[A-Z]*[a-z]*\d*)', site)
    host = host.group('host')
    site_dict = {''.join(site_name.split()): site_name for site_name in os.listdir(SITES_FOLDER)}
    if host in site_dict:
        #file_list = os.listdir(SITES_FOLDER+site_dict[host])
        active_ports_file = glob.glob(SITES_FOLDER+site_dict[host] + '/*active ports*')
        if not active_ports_file:
            print("Not found active ports file site:", site)
            return {'header': '', 'content': ''}
        else:
            wb = load_workbook(active_ports_file[0])
            first_sheet = wb.get_sheet_names()[0]
            worksheet = wb.get_sheet_by_name(first_sheet)
            active_ports = []
            for row in worksheet.iter_rows():
                new_row = [cell.value for cell in row if cell.value ]
                if new_row:
                    active_ports.append(new_row)
            active_ports = {'header': active_ports[0], 'content': active_ports[1::]}
            return active_ports


def get_configuration(old_hostname, ssh):
    with open(SITES_SOFT_CUR+old_hostname+'.cfg') as old_config:
        old_config = old_config.read()
    #print ('gather configuration from:',old_hostname )
    #old_config = send_show_command(ssh, 'show running-config')
    #иначе не работает темлейт
    old_config = old_config.replace('&', 'and')

    new_config = send_show_command(ssh, 'display  current-configuration ')
    #new_config_file = glob.glob(SITES_SOFT_TAR+'{}*'.format(old_hostname))[0]
    #with open(new_config_file) as new_config:
    #    new_config = new_config.read()
    return {'old_config': old_config, 'new_config': new_config}


def save_conf(config, hostname):
    path = './Cisco_conf/{}_cisco.cfg'.format(hostname)
    print('save file:', path)
    with open(path, 'w') as out_file:
        out_file.writelines(config)


def connect_to_device(ip, device_type):
    print ('connected to', ip)
    device = {'device_type': device_type,
              'ip': ip,
              'username': 'hwswap',
              'password': 'hwswap123'
              }
    ssh_connection = netmiko.ConnectHandler(**device)
    #ssh_connection.enable()
    print('connected succesefull')
    return ssh_connection


def send_show_command(ssh_connection, command):
    result = ssh_connection.send_command(command)
    return result


def disconnect(ssh_connection):
    ssh_connection.disconnect()


def get_vlan_list(ssh):
    vlan_list =[]
    print('collect vlan info')
    vlans = send_show_command(ssh, 'show vlan brief')
    template_path = TEMPLATE_CISCO + 'cisco_ios_show_vlan.template'
    with open(template_path) as template:
        re_table = textfsm.TextFSM(template)
        vlan_info = re_table.ParseText(vlans)
    for vlan in vlan_info:
        vlan_list.append(vlan[0])
    return vlan_list


def get_vlan_int_list(ssh):
    vlan_int_list = []
    print('collect vlan int list')
    vlans = send_show_command(ssh, 'show ip interface brief | include lan')
    template_path = TEMPLATE_CISCO + 'cisco_ios_show_ip_int_brief.template'
    with open(template_path) as template:
        re_table = textfsm.TextFSM(template)
        vlan_info = re_table.ParseText(vlans)
    for vlan in vlan_info:
        vlan_int_list.append(vlan[0][4:])
    return vlan_int_list


def parse_vlan_int_conf(ssh, vlan_id):
    vlan_conf = send_show_command(ssh, 'show running-config interface  vlan {}'.format(vlan_id))
    #print (vlan_conf)

    ip = re.search('ip address (?P<ip>[\d.]* [\d.]*)', vlan_conf)
    vrf = re.search('ip vrf forwarding (?P<vrf>[\w\d]*)', vlan_conf)
    xconnect = re.search(' (?P<x>xconnect [\d.]* \d+) encapsulation mpls', vlan_conf)
    if ip:
        ip = ip.group('ip')
    else:
        ip = None
    if vrf:
        vrf = vrf.group('vrf')
    else:
        vrf = None
    if xconnect:
        xconnect = xconnect.group('x')
    else:
        xconnect = None
    return ip, vrf, xconnect


def generate_vlan_table_content(ssh):
    vlan_list = get_vlan_list(ssh)
    vlan_int_list = get_vlan_int_list(ssh)
    content = []
    for vlan in vlan_int_list:
        ip, vrf, xcon = parse_vlan_int_conf(ssh, vlan)
        content.append([vlan, 'yes', ip, vrf, xcon])
    simple_vlan = set(vlan_list).difference(set(vlan_int_list))
    print (simple_vlan)
    return content


def get_int_description_from_cfg_file(hostname):
    with open(SITES_SOFT_CUR+hostname+'.cfg') as old_config:
        old_config = old_config.read()
    with open(TEMPLATE_CISCO + 'cisco_int_des_from_cfg.template') as template:
        re_table = textfsm.TextFSM(template)
        result = re_table.ParseText(old_config)
    return result


def get_media_type(dis_int):
    with open(TEMPLATE_CISCO + 'hw_dis_int.template') as template:
        re_table = textfsm.TextFSM(template)
        result = re_table.ParseText(dis_int)
    return result


def create_interface_table(cisco_ints_description, huawei_ints_description, media_types):
    result = []
    for cisco_int_description in cisco_ints_description:
        for huawei_int_description in huawei_ints_description:
            if cisco_int_description[1] == huawei_int_description[1]:
                for media_type in media_types:
                    if media_type[0] == huawei_int_description[0][2:]:
                        result.append([cisco_int_description[0],
                                       huawei_int_description[0],
                                       cisco_int_description[1],
                                       media_type[1]+' '+media_type[2], '', ''])
    return result


def create_parser():
    parser = argparse.ArgumentParser(description='generate plan')
    parser.add_argument('-collect_config', action="store_true", help ="collect config")
    parser.add_argument('-f', action="store_true", help="full run ip")
    parser.add_argument('-dd', action="store_true", help="full run dd")
    return parser


if __name__ == '__main__':
    parser = create_parser()
    args = parser.parse_args()
    sites_list = get_sites_list()
    sites_information = get_sites_information()
    if args.dd:
        tpl = DocxTemplate(TEMPLATE_DD)
        for site in sites_list:
            context = {}
            print('start working with ', site)
            site_information = sites_information[site]
            cisco_int_description = get_int_description_from_cfg_file(site_information['old_hostname'])
            context = generate_template(tpl, site, site_information)
            context = update_context_from_ssh_data(site_information, context)
            pprint.pprint(context)
            genetate_implementation_design(context, tpl, 'DD')
            context = None
    if args.f:
        tpl = DocxTemplate(TEMPLATE)
        for site in sites_list:
            site_information = sites_information[site]
            print('start working with ', site)
            cisco_int_description = get_int_description_from_cfg_file(site_information['old_hostname'])
            context = generate_template(tpl, site, site_information)
            context = update_context_from_ssh_data(site_information, context)
        save_conf(context['cisco_conf_n'], context['hostname'])
            #genetate_implementation_design(context, tpl, 'ip')
    if args.collect_config:
        for site in sites_list:
            print (site)
            site_information = sites_information[site]
            print (site_information['old_hostname'])
            loopback = site_information['loopback'].split()[0]
            ssh = connect_to_device(loopback, 'cisco_ios')
            config = send_show_command(ssh, 'show running-config')
            save_conf(config, site_information['old_hostname'])
            disconnect(ssh)
